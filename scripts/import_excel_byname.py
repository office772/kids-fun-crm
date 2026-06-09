#!/usr/bin/env python3
"""
Pass 2: Cross-reference skipped Excel rows (no phone) against existing DB parents by name.
If parent name matches → attach child record to that parent.
"""

import openpyxl
import os
import re
import unicodedata
from datetime import datetime
from supabase import create_client

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
EXCEL_PATH   = os.environ.get(
    "EXCEL_PATH",
    "/Users/einatganel/Documents/claude-projects/kids-fun-app/Copy of גבייה תשפ_ו לעינת .xlsx"
)

AREA_MAP = {
    "גני תל אביב": "telaviv",
    "מתן":         "sharon",
    "רשפון":       "sharon",
    "גלי עתלית":   "carmel",
}

def norm_phone(raw):
    if not raw:
        return None
    p = re.sub(r"[\s\-\.\(\)\+]", "", str(raw).replace(".0", "").strip())
    if p.startswith("972"):
        p = "0" + p[3:]
    if not p.startswith("0") and len(p) == 9:
        p = "0" + p
    if not re.match(r"^0\d{8,9}$", p):
        return None
    return p

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("None", "", "-", "לא", "אין", "אין העדפה", "nan") else s

def parse_date(val):
    if not val or str(val) in ("None", ""):
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return None

def parse_dietary(val):
    s = clean(val)
    if not s:
        return None
    if "טבעוני" in s:
        return "טבעוני"
    if "צמחוני" in s:
        return "צמחוני"
    return None

def normalize_name(name):
    """Normalize Hebrew name for fuzzy comparison: strip extra spaces, normalize unicode."""
    if not name:
        return ""
    # NFKC unicode normalization
    n = unicodedata.normalize("NFKC", name.strip())
    # collapse whitespace
    n = re.sub(r"\s+", " ", n)
    return n.lower()

def name_score(a, b):
    """Simple overlap score: how many words from b appear in a (or vice versa)."""
    wa = set(normalize_name(a).split())
    wb = set(normalize_name(b).split())
    if not wa or not wb:
        return 0
    overlap = wa & wb
    # score = overlap / max set size — penalizes partial matches
    return len(overlap) / max(len(wa), len(wb))

def main():
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Load ALL existing parents with name + phone
    all_parents_res = sb.table("parents").select("id,name,phone").execute()
    db_parents = all_parents_res.data
    print(f"הורים קיימים ב-DB: {len(db_parents)}")

    # Build name → parent_id lookup
    name_to_parent = {}
    for p in db_parents:
        if p.get("name"):
            key = normalize_name(p["name"])
            name_to_parent[key] = p["id"]

    stats = {
        "matched": 0,
        "children_created": 0,
        "children_updated": 0,
        "no_match": 0,
        "skipped": 0,
    }

    no_match_rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "הוראות קבע בנקאיות":
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        area = AREA_MAP.get(sheet_name, "unknown")

        for i, row in enumerate(rows[1:], 1):
            d = dict(zip(headers, row))

            p1_phone = norm_phone(d.get("טלפון הורה 1") or d.get("טלפון"))
            p2_phone = norm_phone(d.get("טלפון הורה 2"))

            # Only process rows that were SKIPPED in pass 1 (no valid phone)
            if p1_phone or p2_phone:
                continue

            child_name  = clean(d.get("Name") or d.get("שם הילד") or d.get("שם ילד/ה"))
            parent_name = clean(d.get("שם ההורה")) or clean(d.get("שם הורה 1"))

            if not child_name and not parent_name:
                stats["skipped"] += 1
                continue

            # ── Try exact name match first ──────────────────────────────────
            parent_id = None

            if parent_name:
                key = normalize_name(parent_name)
                if key in name_to_parent:
                    parent_id = name_to_parent[key]

            # ── Fuzzy fallback: find best scoring match ──────────────────────
            if not parent_id and parent_name:
                best_score = 0
                best_id = None
                for p in db_parents:
                    score = name_score(parent_name, p.get("name", ""))
                    if score > best_score:
                        best_score = score
                        best_id = p["id"]
                if best_score >= 0.75:  # require ≥75% word overlap
                    parent_id = best_id

            if not parent_id:
                stats["no_match"] += 1
                no_match_rows.append({
                    "sheet": sheet_name,
                    "row": i,
                    "child": child_name,
                    "parent": parent_name,
                })
                continue

            stats["matched"] += 1

            # ── Upsert child ────────────────────────────────────────────────
            child_id_num = clean(str(d.get("תעודת זהות") or "").replace(".0","").replace("nan",""))
            birth_date   = parse_date(d.get("תאריך לידה"))
            gender       = clean(d.get("מין"))
            city         = clean(d.get("עיר מגורים") or d.get("עיר"))
            school       = clean(d.get("בית חינוכי") or d.get("בית ספר"))
            grade        = clean(d.get("כיתה"))
            allergy_raw  = clean(d.get("מהי סוג הרגישות?") or d.get("האם יש רגישות למזון?") or d.get("האם לילדך יש רגישות למזון?"))
            dietary      = parse_dietary(d.get("צמחוני | טבעוני") or d.get("הילד/ה שלי צמחוני | טבעוני"))
            medical_note = clean(d.get("במידה וקיימת בעיה רפואית צרפו כאן אישור מרופא"))

            if allergy_raw and re.search(r"אין|None|לא|nan", allergy_raw, re.IGNORECASE):
                allergy_raw = None

            # Also update parent city/email if available
            email = clean(d.get("כתובת מייל") or d.get("מייל"))
            parent_patch = {k: v for k, v in {"city": city, "email": email}.items() if v}
            if parent_patch:
                sb.table("parents").update(parent_patch).eq("id", parent_id).execute()

            # Find or create child
            existing_child = None
            if child_id_num and child_id_num not in ("", "None"):
                res = sb.table("children").select("id").eq("id_number", child_id_num).limit(1).execute()
                if res.data:
                    existing_child = res.data[0]

            if not existing_child:
                res = sb.table("children").select("id").eq("parent_id", parent_id).limit(1).execute()
                if res.data:
                    existing_child = res.data[0]

            child_data = {k: v for k, v in {
                "parent_id":     parent_id,
                "name":          child_name or "ילד/ה",
                "birth_date":    birth_date,
                "gender":        gender,
                "area_code":     area,
                "school":        school,
                "grade":         grade,
                "allergies":     allergy_raw,
                "dietary":       dietary,
                "medical_notes": medical_note,
                "id_number":     child_id_num if child_id_num and child_id_num not in ("", "None") else None,
            }.items() if v is not None}

            if existing_child:
                sb.table("children").update(child_data).eq("id", existing_child["id"]).execute()
                stats["children_updated"] += 1
            else:
                try:
                    sb.table("children").insert(child_data).execute()
                    stats["children_created"] += 1
                except Exception as e:
                    print(f"  ⚠️ שגיאת ילד שורה {i} ({sheet_name}): {e}")

    print("\n" + "="*55)
    print("✅ Pass 2 — צליבה לפי שם הסתיימה!")
    print(f"  נמצאו התאמות:     {stats['matched']}")
    print(f"  ילדים נוצרו:      {stats['children_created']}")
    print(f"  ילדים עודכנו:     {stats['children_updated']}")
    print(f"  ללא התאמה:        {stats['no_match']}")
    print(f"  דולגו (ריקים):    {stats['skipped']}")

    if no_match_rows:
        print(f"\n📋 שורות ללא התאמה ({len(no_match_rows)}):")
        for r in no_match_rows[:30]:
            print(f"  [{r['sheet']} #{r['row']}] הורה: {r['parent']} | ילד: {r['child']}")
        if len(no_match_rows) > 30:
            print(f"  ... ועוד {len(no_match_rows)-30} שורות")

if __name__ == "__main__":
    main()
