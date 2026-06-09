#!/usr/bin/env python3
"""
Import Excel data (גבייה תשפ"ו) into Supabase.
Reads credentials from environment variables:
  SUPABASE_URL  + SUPABASE_SERVICE_ROLE_KEY
"""

import openpyxl
import os
import re
import sys
from datetime import datetime
from supabase import create_client

# ─── Config ───────────────────────────────────────────────────────────────────
SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_KEY  = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
EXCEL_PATH    = os.environ.get(
    "EXCEL_PATH",
    "/Users/einatganel/Documents/claude-projects/kids-fun-app/Copy of גבייה תשפ_ו לעינת .xlsx"
)

AREA_MAP = {
    "גני תל אביב": "telaviv",
    "מתן":         "sharon",
    "רשפון":       "sharon",
    "גלי עתלית":   "carmel",
}

# ─── Helpers ──────────────────────────────────────────────────────────────────
def norm_phone(raw):
    if not raw:
        return None
    p = re.sub(r"[\s\-\.\(\)\+]", "", str(raw).replace(".0", "").strip())
    if p.startswith("972"):
        p = "0" + p[3:]
    if not p.startswith("0") and len(p) == 9:
        p = "0" + p
    if not re.match(r"^0\d{8,9}$", p):
        return None
    return p

def parse_date(val):
    if not val or str(val) in ("None", ""):
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return None

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("None", "", "-", "לא", "אין", "אין העדפה", "nan") else s

def parse_dietary(val):
    s = clean(val)
    if not s:
        return None
    if "טבעוני" in s:
        return "טבעוני"
    if "צמחוני" in s:
        return "צמחוני"
    return None

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    stats = {
        "parents_created": 0,
        "parents_updated": 0,
        "children_created": 0,
        "children_updated": 0,
        "skipped": 0,
        "errors": [],
    }

    # Pre-load existing parents (phone → id map)
    existing_raw = sb.table("parents").select("id,phone,name").execute()
    phone_to_id = {}
    for p in existing_raw.data:
        n = norm_phone(p["phone"])
        if n:
            phone_to_id[n] = p["id"]

    print(f"DB הורים קיימים: {len(phone_to_id)}")

    for sheet_name in wb.sheetnames:
        if sheet_name == "הוראות קבע בנקאיות":
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        area = AREA_MAP.get(sheet_name, "unknown")
        print(f"\n📋 {sheet_name} ({len(rows)-1} ילדים, אזור: {area})")
        print(f"   כותרות: {headers}")

        for i, row in enumerate(rows[1:], 1):
            d = dict(zip(headers, row))

            # ── Map column names (flexible — sheets differ slightly) ─────────
            child_name   = clean(d.get("Name") or d.get("שם הילד") or d.get("שם ילד/ה"))
            child_id_num = clean(str(d.get("תעודת זהות") or "").replace(".0","").replace("nan",""))
            birth_date   = parse_date(d.get("תאריך לידה"))
            gender       = clean(d.get("מין"))
            city         = clean(d.get("עיר מגורים") or d.get("עיר"))
            school       = clean(d.get("בית חינוכי") or d.get("בית ספר"))
            grade        = clean(d.get("כיתה"))
            email        = clean(d.get("כתובת מייל") or d.get("מייל") or d.get("אימייל"))
            p1_phone     = norm_phone(d.get("טלפון הורה 1") or d.get("טלפון"))
            p1_name      = clean(d.get("שם הורה 1") or d.get("שם ההורה"))
            p2_phone     = norm_phone(d.get("טלפון הורה 2"))
            p2_name      = clean(d.get("שם הורה 2"))
            parent_name  = clean(d.get("שם ההורה")) or p1_name
            allergy_raw  = clean(d.get("מהי סוג הרגישות?") or d.get("האם יש רגישות למזון?") or d.get("האם לילדך יש רגישות למזון?") or d.get("אלרגיה"))
            dietary      = parse_dietary(d.get("צמחוני | טבעוני") or d.get("הילד/ה שלי צמחוני | טבעוני") or d.get("תזונה"))
            medical_note = clean(d.get("במידה וקיימת בעיה רפואית צרפו כאן אישור מרופא") or d.get("בעיה רפואית"))
            notes        = clean(d.get("הוספת הערה / בקשה") or d.get("הערות"))
            parent_id_num = clean(str(d.get("ת״ז הורה") or "").replace(".0","").replace("nan",""))

            # Normalize allergy
            if allergy_raw and re.search(r"אין|None|לא|nan", allergy_raw, re.IGNORECASE):
                allergy_raw = None

            # Skip rows with no useful data
            if not p1_phone and not child_name:
                stats["skipped"] += 1
                continue

            # ── Find or create parent ───────────────────────────────────────
            parent_id = None

            if p1_phone and p1_phone in phone_to_id:
                parent_id = phone_to_id[p1_phone]
            elif p2_phone and p2_phone in phone_to_id:
                parent_id = phone_to_id[p2_phone]

            if parent_id:
                patch = {}
                if city:      patch["city"]          = city
                if email:     patch["email"]         = email
                if p2_name:   patch["parent2_name"]  = p2_name
                if p2_phone:  patch["parent2_phone"] = p2_phone
                if patch:
                    sb.table("parents").update(patch).eq("id", parent_id).execute()
                    stats["parents_updated"] += 1
            else:
                phone_to_use = p1_phone or f"xl_{sheet_name[:4]}_{i}"
                try:
                    insert_data = {k: v for k, v in {
                        "name":          parent_name or child_name or "לא ידוע",
                        "phone":         phone_to_use,
                        "email":         email,
                        "city":          city,
                        "parent2_name":  p2_name,
                        "parent2_phone": p2_phone,
                        "id_number":     parent_id_num if parent_id_num and parent_id_num not in ("", "None") else None,
                        "sync_source":   "excel_import",
                        "contact_type":  "parent",
                        "notes":         notes,
                    }.items() if v is not None}
                    res = sb.table("parents").insert(insert_data).execute()
                    parent_id = res.data[0]["id"]
                    phone_to_id[phone_to_use] = parent_id
                    if p2_phone and p2_phone not in phone_to_id:
                        phone_to_id[p2_phone] = parent_id
                    stats["parents_created"] += 1
                except Exception as e:
                    stats["errors"].append(f"Parent error row {i} ({sheet_name}): {e}")
                    stats["skipped"] += 1
                    continue

            if not parent_id:
                stats["skipped"] += 1
                continue

            # ── Find or create child ────────────────────────────────────────
            existing_child = None
            if child_id_num and child_id_num not in ("", "None"):
                res = sb.table("children").select("id").eq("id_number", child_id_num).limit(1).execute()
                if res.data:
                    existing_child = res.data[0]

            if not existing_child and child_name:
                res = sb.table("children").select("id").eq("parent_id", parent_id).limit(1).execute()
                if res.data:
                    existing_child = res.data[0]

            child_data = {k: v for k, v in {
                "parent_id":      parent_id,
                "name":           child_name or "ילד/ה",
                "birth_date":     birth_date,
                "gender":         gender,
                "area_code":      area,
                "school":         school,
                "grade":          grade,
                "allergies":      allergy_raw,
                "dietary":        dietary,
                "medical_notes":  medical_note,
                "id_number":      child_id_num if child_id_num and child_id_num not in ("", "None") else None,
            }.items() if v is not None}

            if existing_child:
                sb.table("children").update(child_data).eq("id", existing_child["id"]).execute()
                stats["children_updated"] += 1
            else:
                try:
                    sb.table("children").insert(child_data).execute()
                    stats["children_created"] += 1
                except Exception as e:
                    stats["errors"].append(f"Child error row {i} ({sheet_name}): {e}")

            if i % 100 == 0:
                print(f"  ... {i} שורות עובדו")

    print("\n" + "="*50)
    print("✅ ייבוא הסתיים!")
    print(f"  הורים נוצרו:    {stats['parents_created']}")
    print(f"  הורים עודכנו:   {stats['parents_updated']}")
    print(f"  ילדים נוצרו:    {stats['children_created']}")
    print(f"  ילדים עודכנו:   {stats['children_updated']}")
    print(f"  דולגו:          {stats['skipped']}")
    if stats["errors"]:
        print(f"\n⚠️  שגיאות ({len(stats['errors'])}):")
        for e in stats["errors"][:20]:
            print(f"  {e}")

if __name__ == "__main__":
    main()
